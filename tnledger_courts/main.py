from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from webdriver_manager.chrome import ChromeDriverManager
import datefinder, random
from datetime import date
import PySimpleGUI as sg
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account

entered_dates=''
today = str(date.today()).split('-')
theme_name_list = sg.theme_list()

while True:
    sg.theme(theme_name_list[random.randint(0, len(theme_name_list))])
    #define layout
    layout=[
        [sg.Text('Enter the date',size=(20, 1), font='Ubuntu',justification='left')],
            [sg.Input(key='from', size=(20,1)), sg.CalendarButton('Calendar',font="Ubuntu",  target='from', default_date_m_d_y=(int(today[1]),int(today[2]),int(today[0])), )],
            [sg.Button('OK', font=('Ubuntu',12)),sg.Button('CANCEL', font=('Ubuntu',12))]]
    #Define Window
    win =sg.Window('Montgomery',layout)
    #Read  values entered by user
    e,v=win.read()
    con = False 
    print(e,v)
    if e == None or e == "CANCEL":
        print('ham')
        print("exit")
        win.close()
        con = True
        break
    else:
        if  v['from'] == None or v['from'] == '':
            print('Enter the date correctly')
            
            win.close()
            continue
        else:
            entered_dates = f"{v['from'].split(' ')[0].split('-')[1]}/{v['from'].split(' ')[0].split('-')[2]}/{v['from'].split(' ')[0].split('-')[0]}"
            win.close()
            break
print(entered_dates)

def remove_values_from_list(the_list, val):
    return [value for value in the_list if value != val]
if con:
    pass
else:
    driver1 = webdriver.Chrome(ChromeDriverManager().install())
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get(f"https://www.tnledger.com/Notices.aspx?noticesDate={entered_dates}")
    driver.maximize_window()
    driver.minimize_window()
    driver.implicitly_wait(30)
    dates=entered_dates
    court_rows=driver.find_elements(By.XPATH, '//table[@id="ContentPane_CourtGridView"]/tbody/tr')
    del court_rows[0]

    date=driver.current_url[driver.current_url.find('=')+1:]
    cheatham_county_data=[]
    montgomery_county_data=[]
    davidson_county_data=[]
    for i in court_rows:
        individual_rows=i.find_elements(By.TAG_NAME, 'td')
        each_row_data=[]
        for j in (individual_rows):
            if j.text=="View":
                current_data_cheatham=[]
                current_data_montgomery=[]
                current_data_davidson=[]
                current_link=(j.find_element(By.TAG_NAME, 'a').get_attribute('href'))
                link_id=current_link[current_link.find('(')+1:current_link.find(')')].split(',')[0].replace("'","") 
                
                driver1.get(f"https://www.tnledger.com/Search/Details/ViewNotice.aspx?id={link_id}&date={date}")
    #             driver1.minimize_window()
                driver1.implicitly_wait(30)
                
                page_data=driver1.find_element(By.XPATH, '//div[@id="record-details"]')
                extracted_text=(page_data.text.splitlines())
                probate_status=False
                cheatham_status=False
                montgomery_status=False
                davidson_status=False
                for i in extracted_text:
                    if 'PROBATE' in i:
                        probate_status=True
                    if 'CHEATHAM COUNTY' in i:
                        cheatham_status=True
                    if 'MONTGOMERY COUNTY' in i:
                        montgomery_status=True
                    if 'DAVIDSON COUNTY' in i:
                        davidson_status=True
                city='Not Available'
                executors='Not Available'
                attorney='Not Available'
                estate_of='Not Available'
                county_name='Not Available'
                    
                if probate_status==True and davidson_status==True:
                    data=page_data.text
                    print('DAVIDSON COUNTY whole data going to csv.')
                    current_data_davidson.append(data)
                    davidson_county_data.append(current_data_davidson)
                if probate_status==True and cheatham_status==True:
                    for i in extracted_text:
                        if 'CHEATHAM COUNTY' in i:
                            splited=i.split()
                            county_name=splited[splited.index('COUNTY')-1]
                            
                            print(county_name)
                            city=extracted_text[extracted_text.index(i)+1]
                            print(city)
                            
                        if 'Estate of' in i:
                            estate_of=(i.replace('Estate of','').replace(', Deceased', '').strip())
                            print(estate_of)
                            
                        if 'Attorney' in i:
                            attorney=(i.replace('Attorney','').strip())
                            attorney=attorney.replace(':','').strip()
                            
                            print(attorney.strip())
                        if 'Executor'in i:
                            space=1
                            executors=''
                            while True:
                                executors=executors+'|'+(extracted_text[extracted_text.index(i)-space]).replace(' and ','')
                                space=space+1
                                if len((list(datefinder.find_dates(extracted_text[extracted_text.index(i)-space])))) >0:
                                    break
                            executors=executors.strip()
                            print(executors)
                            
                        if 'Administrator' in i:
                            space=1
                            executors=''
                            while True:
                                executors=executors+'|'+(extracted_text[extracted_text.index(i)-space]).replace(' and ','')
                                space=space+1
                                if len((list(datefinder.find_dates(extracted_text[extracted_text.index(i)-space])))) >0:
                                    break
                            executors=executors.strip()
                            print(executors)
                            
                        if 'Executrix' in i:
                            space=1
                            executors=''
                            while True:
                                executors=executors+'|'+(extracted_text[extracted_text.index(i)-space]).replace(' and ','')
                                space=space+1
                                if len((list(datefinder.find_dates(extracted_text[extracted_text.index(i)-space])))) >0:
                                    break
                            executors=executors.strip()
                            print(executors)
                            
                            
                        if 'Administratrix' in i:
                            space=1
                            executors=''
                            while True:
                                executors=executors+'|'+(extracted_text[extracted_text.index(i)-space]).replace(' and ','')
                                space=space+1
                                if len((list(datefinder.find_dates(extracted_text[extracted_text.index(i)-space])))) >0:
                                    break
                            executors=executors.strip()
                            print(executors)
                            
                        if 'Executris' in i:
                            space=1
                            executors=''
                            while True:
                                executors=executors+'|'+(extracted_text[extracted_text.index(i)-space]).replace(' and ','')
                                space=space+1
                                if len((list(datefinder.find_dates(extracted_text[extracted_text.index(i)-space])))) >0:
                                    break
                            executors=executors.strip()
                            print(executors)
                    print([county_name,city,estate_of,executors,attorney])  
                    if executors.strip()!='':
                        if executors[0] == '|':
                            executors = executors[1:]
                    cheatham_county_data.append([city,estate_of,executors,attorney])
                    
                if probate_status==True and montgomery_status==True:
                    for i in extracted_text:
                        if 'MONTGOMERY COUNTY' in i:
                            county_name='MONTGOMERY'
                            current_data_montgomery.append(county_name)
                        if 'estate of' and '(deceased)' in i.lower():
                            estate_of=i.replace('ESTATE OF','').replace('(Deceased)', '').strip()
                        if 'Attorney' in i:
                            print(i.split(':')[-1].strip())
                            attorney=i.split(':')[-1].strip()
                            current_data_montgomery.append(attorney)
                            current_pos=extracted_text.index(i)
                            space=1
                            executors=''
                            while True:
                                if len(list(datefinder.find_dates(extracted_text[current_pos-space])))>0:
                                       break
                                executors=executors+'|'+(extracted_text[current_pos-space])
                                space=space+1
                            
                        if 'deputy' in i.lower():
                            city=(extracted_text[extracted_text.index(i)-1])
                    print([county_name,city,estate_of,executors,attorney])
                    if executors.strip()!='':
                        if executors[0] == '|':
                            executors = executors[1:]
                        executors = executors.split('-')[0]
                    montgomery_county_data.append([city,estate_of,executors,attorney])
    driver1.close()

    cheatham_county_data=remove_values_from_list(cheatham_county_data,[])
    montgomery_county_data=remove_values_from_list(montgomery_county_data,[])
    
    ##########################################################################################################
#     bold_font = Font(bold=True)
#     bold_font = Font(bold=True)
#     center_aligned_text = Alignment(horizontal="center")
#     double_border_side = Side(border_style="double")
#     square_border = Border(top=double_border_side,
#                     right=double_border_side,
#                     bottom=double_border_side,
#                     left=double_border_side)
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title='Cheatham'
#     sheet["A1"] = "County Name"
#     sheet["A1"].font = bold_font
#     sheet["A1"].alignment = center_aligned_text
#     sheet["A1"].border = square_border
#     sheet["B1"] = "City"
#     sheet["B1"].font = bold_font
#     sheet["B1"].alignment = center_aligned_text
#     sheet["B1"].border = square_border
#     sheet["C1"] = "Estate of"
#     sheet["C1"].font = bold_font
#     sheet["C1"].alignment = center_aligned_text
#     sheet["C1"].border = square_border
#     sheet["D1"] = "Executors"
#     sheet["D1"].font = bold_font
#     sheet["D1"].alignment = center_aligned_text
#     sheet["D1"].border = square_border
#     sheet["E1"] = "Attorney"
#     sheet["E1"].font = bold_font
#     sheet["E1"].alignment = center_aligned_text
#     sheet["E1"].border = square_border

#     dim_holder = DimensionHolder(worksheet=sheet)

#     for col in range(sheet.min_column, sheet.max_column + 1):
#         dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

#     sheet.column_dimensions = dim_holder
#     for p,q in enumerate(cheatham_county_data):
#         sheet[f"A{p+2}"]=q[0]           
#         sheet[f"B{p+2}"]=q[1]
#         sheet[f"C{p+2}"]=q[2]
#         sheet[f"D{p+2}"]=q[3]
#         sheet[f"E{p+2}"]=q[4]
#     sheet=workbook.create_sheet("Sheet_A")
#     sheet.title='Montgomery'
#     sheet["A1"] = "County Name"
#     sheet["A1"].font = bold_font
#     sheet["A1"].alignment = center_aligned_text
#     sheet["A1"].border = square_border
#     sheet["B1"] = "City"
#     sheet["B1"].font = bold_font
#     sheet["B1"].alignment = center_aligned_text
#     sheet["B1"].border = square_border
#     sheet["C1"] = "Estate Of"
#     sheet["C1"].font = bold_font
#     sheet["C1"].alignment = center_aligned_text
#     sheet["C1"].border = square_border
#     sheet["D1"] = "Executors"
#     sheet["D1"].font = bold_font
#     sheet["D1"].alignment = center_aligned_text
#     sheet["D1"].border = square_border
#     sheet["E1"] = "Attorney"
#     sheet["E1"].font = bold_font
#     sheet["E1"].alignment = center_aligned_text
#     sheet["E1"].border = square_border

#     dim_holder = DimensionHolder(worksheet=sheet)

#     for col in range(sheet.min_column, sheet.max_column + 1):
#         dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

#     sheet.column_dimensions = dim_holder
#     for p,q in enumerate(montgomery_county_data):
#         sheet[f"A{p+2}"]=q[0]           
#         sheet[f"B{p+2}"]=q[1]
#         sheet[f"C{p+2}"]=q[2]
#         sheet[f"D{p+2}"]=q[3]
#         sheet[f"E{p+2}"]=q[4]
#     sheet=workbook.create_sheet("Sheet_B")
#     sheet.title='Davidson'
#     sheet["A1"] = "Whole data"
#     sheet["A1"].font = bold_font
#     sheet["A1"].alignment = center_aligned_text
#     sheet["A1"].border = square_border
#     dim_holder = DimensionHolder(worksheet=sheet)

#     for col in range(sheet.min_column, sheet.max_column + 1):
#         dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

#     sheet.column_dimensions = dim_holder
#     for p,q in enumerate(davidson_county_data):
#         sheet[f"A{p+2}"]=q[0]           
        
#     workbook.save(f"tnledger_court_notices[{dates.replace('/','-')}].xlsx")
    
    ###################################################################################################################
    
    SERVICE_ACCOUNT_FILE = 'keys.json'
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    creds = None
    creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    SAMPLE_SPREADSHEET_ID = '1KIcOYi8gB5ZvHEXRGopqineBlbDx1sH7fAxRyCPROB0'

    try:
        service = build('sheets', 'v4', credentials=creds)

        sheet = service.spreadsheets()
        
        montgomery_county_data = [list(t) for t in set(tuple(element) for element in montgomery_county_data)]
        montgomery_county_data.insert(0, [f'SELECTED DATE   =>   {entered_dates}'])
        montgomery_county_data.insert(0, [' - '])
        montgomery_county_data.insert(2, [' - '])
        
        cheatham_county_data = [list(t) for t in set(tuple(element) for element in cheatham_county_data)]
        cheatham_county_data.insert(0, [f'SELECTED DATE   =>   {entered_dates}'])
        cheatham_county_data.insert(0, [' - '])
        cheatham_county_data.insert(2, [' - '])
        
        davidson_county_data = [list(t) for t in set(tuple(element) for element in davidson_county_data)]
        davidson_county_data.insert(0, [f'SELECTED DATE   =>   {entered_dates}'])
        davidson_county_data.insert(0, [' - '])
        davidson_county_data.insert(2, [' - '])

        
        request = sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                    range="Cheatham!A2", valueInputOption="USER_ENTERED", body={"values":cheatham_county_data}).execute()

        request = sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                    range="Montgomery!A2", valueInputOption="USER_ENTERED", body={"values":montgomery_county_data}).execute()

        request = sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                    range="Davidson!A2", valueInputOption="USER_ENTERED", body={"values":davidson_county_data}).execute()

    except HttpError as err:
        print(err)

    ###################################################################################################################
    
    driver.close()
    print('Hey Angel, scraping is done. You can now close this terminal')
